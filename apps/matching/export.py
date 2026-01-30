"""Export functionality for match results."""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .services import get_match_run_results


def export_match_run_xlsx(match_run):
    """
    Export match run results to XLSX format.

    Args:
        match_run: MatchRun object

    Returns:
        Bytes content of XLSX file
    """
    results = get_match_run_results(match_run)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Match Results {match_run.id}"

    # Define headers
    headers = [
        "Mentor Name",
        "Mentor Email",
        "Mentor Organization",
        "Mentee Name",
        "Mentee Email",
        "Mentee Organization",
        "Match Percentage",
        "Ambiguity Flag",
        "Ambiguity Reason",
        "Exception Flag",
        "Exception Type",
        "Exception Reason",
        "Manual Override",
        "Override Reason",
    ]

    # Write headers with styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_num, result in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=result["mentor_name"])
        ws.cell(row=row_num, column=2, value=result["mentor_email"])
        ws.cell(row=row_num, column=3, value=result["mentor_org"])
        ws.cell(row=row_num, column=4, value=result["mentee_name"])
        ws.cell(row=row_num, column=5, value=result["mentee_email"])
        ws.cell(row=row_num, column=6, value=result["mentee_org"])
        ws.cell(row=row_num, column=7, value=result["match_percent"])
        ws.cell(row=row_num, column=8, value=result["ambiguity_flag"])
        ws.cell(row=row_num, column=9, value=result["ambiguity_reason"])
        ws.cell(row=row_num, column=10, value=result["exception_flag"])
        ws.cell(row=row_num, column=11, value=result["exception_type"])
        ws.cell(row=row_num, column=12, value=result["exception_reason"])
        ws.cell(row=row_num, column=13, value=result["is_manual_override"])
        ws.cell(row=row_num, column=14, value=result["override_reason"])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add metadata
    metadata_row = len(results) + 3
    ws.cell(row=metadata_row, column=1, value=f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=metadata_row + 1, column=1, value=f"Match Run ID: {match_run.id}")
    ws.cell(row=metadata_row + 2, column=1, value=f"Cohort: {match_run.cohort.name}")
    ws.cell(row=metadata_row + 3, column=1, value=f"Mode: {match_run.get_mode_display()}")

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()